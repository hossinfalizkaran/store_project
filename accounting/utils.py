import os
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from django.http import FileResponse
import subprocess
import zipfile
import jdatetime
import pytz
from datetime import datetime
from .models import Sanad, SanadDetail, Perinf, Pergrp
from django.core.paginator import Paginator
from django.db import models
import shutil
import re
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from .models import Settings
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Users
import hashlib
from django.db import connections

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.utils import timezone
from decimal import Decimal
import json
from datetime import datetime

from .models import (
    Perinf, Goodinf, Sanad, SanadDetail, Kardex,
    FactFo, FactFoDetail, Kharid, ChequesRecieve, ChequePay,
    Chequerecievelog, ChequepayLog, GetRecieve, Stores
)


# ========================================
# توابع مدیریت تنظیمات UI
# ========================================

def get_setting(user_id, category, code, default_value=None):
    """
    دریافت تنظیم خاص از دیتابیس با اولویت‌بندی:
    1. تنظیمات شخصی کاربر
    2. تنظیمات عمومی سیستم
    3. مقدار پیش‌فرض
    
    Args:
        user_id (int): شناسه کاربر
        category (int): دسته تنظیمات
        code (int): کد تنظیمات
        default_value: مقدار پیش‌فرض در صورت عدم وجود
    
    Returns:
        str: مقدار تنظیمات
    """
    try:
        # ابتدا جستجوی تنظیمات شخصی کاربر با استفاده از values()
        setting_data = Settings.objects.filter(
            usercode=user_id,
            managetype=1,  # تنظیمات شخصی
            category=category,
            code=code
        ).values('comment').first()
        
        if setting_data:
            return setting_data['comment']
    except Exception:
        pass
    
    try:
        # در صورت عدم وجود، جستجوی تنظیمات عمومی سیستم
        setting_data = Settings.objects.filter(
            usercode=0,
            managetype=0,  # تنظیمات عمومی
            category=category,
            code=code
        ).values('comment').first()
        
        if setting_data:
            return setting_data['comment']
    except Exception:
        pass
    
    # استفاده از مقدار پیش‌فرض
    return default_value


def set_user_setting(user_id, category, code, value):
    """
    ذخیره تنظیم شخصی برای کاربر
    
    Args:
        user_id (int): شناسه کاربر
        category (int): دسته تنظیمات
        code (int): کد تنظیمات
        value (str): مقدار تنظیمات
    """
    setting, created = Settings.objects.get_or_create(
        usercode=user_id,
        managetype=1,  # تنظیمات شخصی
        category=category,
        code=code,
        defaults={'comment': value}
    )
    if not created:
        setting.comment = value
        setting.save()
    return setting


def get_boolean_setting(user_id, category, code, default_value=False):
    """
    دریافت تنظیم boolean از دیتابیس
    
    Args:
        user_id (int): شناسه کاربر
        category (int): دسته تنظیمات
        code (int): کد تنظیمات
        default_value (bool): مقدار پیش‌فرض
    
    Returns:
        bool: مقدار تنظیمات
    """
    try:
        # ابتدا جستجوی تنظیمات شخصی کاربر با استفاده از values()
        setting_data = Settings.objects.filter(
            usercode=user_id,
            managetype=1,  # تنظیمات شخصی
            category=category,
            code=code
        ).values('selected').first()
        
        if setting_data:
            return setting_data['selected']
    except Exception:
        pass
    
    try:
        # در صورت عدم وجود، جستجوی تنظیمات عمومی سیستم
        setting_data = Settings.objects.filter(
            usercode=0,
            managetype=0,  # تنظیمات عمومی
            category=category,
            code=code
        ).values('selected').first()
        
        if setting_data:
            return setting_data['selected']
    except Exception:
        pass
    
    # استفاده از مقدار پیش‌فرض
    return default_value


def set_boolean_setting(user_id, category, code, value):
    """
    ذخیره تنظیم boolean برای کاربر
    
    Args:
        user_id (int): شناسه کاربر
        category (int): دسته تنظیمات
        code (int): کد تنظیمات
        value (bool): مقدار تنظیمات
    """
    setting, created = Settings.objects.get_or_create(
        usercode=user_id,
        managetype=1,  # تنظیمات شخصی
        category=category,
        code=code,
        defaults={'selected': value}
    )
    if not created:
        setting.selected = value
        setting.save()
    return setting


# ========================================
# تنظیمات خاص برای لیست‌های مختلف
# ========================================

def get_person_list_settings(user_id):
    """
    دریافت تنظیمات لیست اشخاص
    
    Returns:
        dict: شامل sort_key, visible_columns, show_inactive
    """
    return {
        'sort_key': get_setting(user_id, 3, 1, 'code'),  # مرتب‌سازی
        'visible_columns': get_setting(user_id, 3, 2, 'code,name,lname,tel1,mobile,credit').split(','),  # ستون‌های قابل نمایش
        'show_inactive': get_boolean_setting(user_id, 3, 3, False),  # نمایش اشخاص غیرفعال
        'page_size': int(get_setting(user_id, 3, 4, '20')),  # تعداد آیتم در صفحه
    }


def get_sanad_list_settings(user_id):
    """
    دریافت تنظیمات لیست اسناد
    
    Returns:
        dict: شامل sort_key, visible_columns, show_cancelled
    """
    return {
        'sort_key': get_setting(user_id, 3, 10, 'date'),  # مرتب‌سازی
        'visible_columns': get_setting(user_id, 3, 11, 'code,date,sharh,status').split(','),  # ستون‌های قابل نمایش
        'show_cancelled': get_boolean_setting(user_id, 3, 12, False),  # نمایش اسناد لغو شده
        'page_size': int(get_setting(user_id, 3, 13, '20')),  # تعداد آیتم در صفحه
    }


def get_sales_list_settings(user_id):
    """
    دریافت تنظیمات لیست فاکتورهای فروش
    
    Returns:
        dict: شامل sort_key, visible_columns, show_cancelled
    """
    return {
        'sort_key': get_setting(user_id, 3, 20, 'tarikh'),  # مرتب‌سازی
        'visible_columns': get_setting(user_id, 3, 21, 'code,tarikh,shakhs_code,mablagh_factor,status').split(','),  # ستون‌های قابل نمایش
        'show_cancelled': get_boolean_setting(user_id, 3, 22, False),  # نمایش فاکتورهای لغو شده
        'page_size': int(get_setting(user_id, 3, 23, '20')),  # تعداد آیتم در صفحه
    }


def get_purchase_list_settings(user_id):
    """
    دریافت تنظیمات لیست فاکتورهای خرید
    
    Returns:
        dict: شامل sort_key, visible_columns, show_cancelled
    """
    return {
        'sort_key': get_setting(user_id, 3, 30, 'tarikh'),  # مرتب‌سازی
        'visible_columns': get_setting(user_id, 3, 31, 'code,tarikh,shakhs_code,mablagh_factor,status').split(','),  # ستون‌های قابل نمایش
        'show_cancelled': get_boolean_setting(user_id, 3, 32, False),  # نمایش فاکتورهای لغو شده
        'page_size': int(get_setting(user_id, 3, 33, '20')),  # تعداد آیتم در صفحه
    }


def get_good_list_settings(user_id):
    """
    دریافت تنظیمات لیست کالاها
    
    Returns:
        dict: شامل sort_key, visible_columns, show_inactive
    """
    return {
        'sort_key': get_setting(user_id, 3, 40, 'name'),  # مرتب‌سازی
        'visible_columns': get_setting(user_id, 3, 41, 'code,name,unit,mogodi,price').split(','),  # ستون‌های قابل نمایش
        'show_inactive': get_boolean_setting(user_id, 3, 42, False),  # نمایش کالاهای غیرفعال
        'page_size': int(get_setting(user_id, 3, 43, '20')),  # تعداد آیتم در صفحه
    }


# ========================================
# توابع کمکی برای پردازش تنظیمات
# ========================================

def parse_sort_key(sort_key):
    """
    پردازش کلید مرتب‌سازی و تبدیل به فیلدهای قابل استفاده در order_by
    
    Args:
        sort_key (str): کلید مرتب‌سازی (مثل 'lname,name' یا 'code')
    
    Returns:
        list: لیست فیلدها برای order_by
    """
    if not sort_key:
        return ['code']
    
    # تبدیل فیلدهای فارسی به انگلیسی
    field_mapping = {
        'code': 'code',
        'name': 'name',
        'lname': 'lname',
        'fullname': 'fullname',
        'tel1': 'tel1',
        'mobile': 'mobile',
        'credit': 'credit',
        'date': 'tarikh',
        'sharh': 'sharh',
        'status': 'status',
        'mablagh_factor': 'mablagh_factor',
        'shakhs_code': 'shakhs_code__name',
        'unit': 'unit__name',
        'mogodi': 'mogodi',
        'price': 'price',
    }
    
    fields = sort_key.split(',')
    result = []
    
    for field in fields:
        field = field.strip()
        if field in field_mapping:
            result.append(field_mapping[field])
        else:
            result.append(field)
    
    return result if result else ['code']


def get_column_display_name(column_name):
    """
    دریافت نام نمایشی ستون
    
    Args:
        column_name (str): نام ستون
    
    Returns:
        str: نام نمایشی ستون
    """
    display_names = {
        'code': 'کد',
        'name': 'نام',
        'lname': 'نام خانوادگی',
        'fullname': 'نام کامل',
        'tel1': 'تلفن',
        'mobile': 'موبایل',
        'credit': 'اعتبار',
        'date': 'تاریخ',
        'tarikh': 'تاریخ',
        'sharh': 'شرح',
        'status': 'وضعیت',
        'mablagh_factor': 'مبلغ کل',
        'shakhs_code': 'مشتری/تامین‌کننده',
        'shakhs_code__name': 'مشتری/تامین‌کننده',
        'unit': 'واحد',
        'unit__name': 'واحد',
        'mogodi': 'موجودی',
        'price': 'قیمت',
        'usercreated': 'کاربر ایجاد کننده',
        'createddate': 'تاریخ ایجاد',
        'createdtime': 'ساعت ایجاد',
    }
    
    return display_names.get(column_name, column_name)


def get_column_value(obj, column_name):
    """
    دریافت مقدار ستون از آبجکت
    
    Args:
        obj: آبجکت مدل
        column_name (str): نام ستون
    
    Returns:
        str: مقدار ستون
    """
    try:
        # برای فیلدهای مربوطه (مثل shakhs_code__name)
        if '__' in column_name:
            parts = column_name.split('__')
            value = obj
            for part in parts:
                value = getattr(value, part)
            return value
        else:
            return getattr(obj, column_name)
    except (AttributeError, TypeError):
        return ''


def format_column_value(value, column_name):
    """
    فرمت کردن مقدار ستون برای نمایش
    
    Args:
        value: مقدار ستون
        column_name (str): نام ستون
    
    Returns:
        str: مقدار فرمت شده
    """
    if value is None:
        return '-'
    
    # فرمت کردن بر اساس نوع ستون
    if column_name in ['mablagh_factor', 'price', 'credit']:
        try:
            return f"{float(value):,.0f} ریال"
        except (ValueError, TypeError):
            return str(value)
    elif column_name in ['mogodi']:
        try:
            return f"{float(value):,.2f}"
        except (ValueError, TypeError):
            return str(value)
    elif column_name in ['status']:
        status_map = {
            1: 'فعال',
            0: 'غیرفعال',
            2: 'لغو شده'
        }
        return status_map.get(value, str(value))
    else:
        return str(value)


# ========================================
# توابع اصلی برای استفاده در View ها
# ========================================

def apply_list_settings(queryset, settings_dict, model_name):
    """
    اعمال تنظیمات لیست روی queryset
    
    Args:
        queryset: queryset اصلی
        settings_dict (dict): تنظیمات لیست
        model_name (str): نام مدل برای فیلتر کردن
    
    Returns:
        queryset: queryset با تنظیمات اعمال شده
    """
    # اعمال فیلترهای boolean
    if model_name == 'Perinf' and not settings_dict.get('show_inactive', False):
        queryset = queryset.filter(status=1)
    elif model_name == 'Sanad' and not settings_dict.get('show_cancelled', False):
        queryset = queryset.filter(status=1)
    elif model_name == 'FactFo' and not settings_dict.get('show_cancelled', False):
        queryset = queryset.filter(status=1)
    elif model_name == 'Kharid' and not settings_dict.get('show_cancelled', False):
        queryset = queryset.filter(status=1)
    elif model_name == 'Goodinf' and not settings_dict.get('show_inactive', False):
        queryset = queryset.filter(status=1)
    
    # اعمال مرتب‌سازی
    sort_fields = parse_sort_key(settings_dict.get('sort_key', 'code'))
    return queryset.order_by(*sort_fields)


def get_list_context(queryset, settings_dict, request):
    """
    ایجاد context برای لیست‌ها
    
    Args:
        queryset: queryset اصلی
        settings_dict (dict): تنظیمات لیست
        request: درخواست HTTP
    
    Returns:
        dict: context برای template
    """
    # اعمال تنظیمات روی queryset
    queryset = apply_list_settings(queryset, settings_dict, queryset.model.__name__)
    
    # صفحه‌بندی
    page_size = settings_dict.get('page_size', 20)
    paginator = Paginator(queryset, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return {
        'object_list': page_obj,
        'settings': settings_dict,
        'visible_columns': settings_dict.get('visible_columns', []),
        'sort_key': settings_dict.get('sort_key', 'code'),
        'page_size': page_size,
    }


# ========================================
# توابع برای ذخیره تنظیمات کاربر
# ========================================

def save_user_list_settings(user_id, category, code, setting_type, value):
    """
    ذخیره تنظیمات لیست برای کاربر
    
    Args:
        user_id (int): شناسه کاربر
        category (int): دسته تنظیمات
        code (int): کد تنظیمات
        setting_type (str): نوع تنظیم ('sort', 'columns', 'boolean')
        value: مقدار تنظیمات
    """
    if setting_type == 'boolean':
        set_boolean_setting(user_id, category, code, value)
    else:
        set_user_setting(user_id, category, code, str(value))


# ========================================
# توابع موجود (حفظ شده)
# ========================================

def export_users_to_excel(request):
    """صادر کردن لیست کاربران به فایل اکسل"""
    try:
        # دریافت تمام کاربران
        users = Users.objects.all()
        
        # ایجاد فایل اکسل
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "لیست کاربران"
        
        # تنظیم هدرها
        headers = ['کد', 'نام کاربری', 'نام', 'نام خانوادگی', 'ایمیل', 'وضعیت']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # اضافه کردن داده‌ها
        for row, user in enumerate(users, 2):
            ws.cell(row=row, column=1).value = user.code
            ws.cell(row=row, column=2).value = user.username
            ws.cell(row=row, column=3).value = user.name
            ws.cell(row=row, column=4).value = user.lname
            ws.cell(row=row, column=5).value = user.email
            ws.cell(row=row, column=6).value = 'فعال' if user.status else 'غیرفعال'
        
        # تنظیم عرض ستون‌ها
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # ذخیره فایل
        filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, filename)
        wb.save(filepath)
        
        # ارسال فایل
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # حذف فایل موقت
        os.remove(filepath)
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def test_password_formats():
    """تست فرمت‌های مختلف رمز عبور"""
    test_passwords = [
        "123456",
        "password",
        "admin123",
        "qwerty",
        "letmein",
        "welcome",
        "monkey",
        "dragon",
        "master",
        "hello"
    ]
    
    results = []
    for password in test_passwords:
        # MD5
        md5_hash = hashlib.md5(password.encode()).hexdigest()
        
        # SHA1
        sha1_hash = hashlib.sha1(password.encode()).hexdigest()
        
        # SHA256
        sha256_hash = hashlib.sha256(password.encode()).hexdigest()
        
        results.append({
            'password': password,
            'md5': md5_hash,
            'sha1': sha1_hash,
            'sha256': sha256_hash
        })
    
    return results


def debug_passwords():
    """دیباگ رمزهای عبور کاربران"""
    users = Users.objects.all()
    debug_info = []
    
    for user in users:
        # بررسی فرمت‌های مختلف
        password_info = {
            'user_id': user.code,
            'username': user.username,
            'original_password': user.password,
            'formats': {}
        }
        
        # تست فرمت‌های مختلف
        test_passwords = ['123456', 'admin', 'password']
        for test_pwd in test_passwords:
            # MD5
            md5_hash = hashlib.md5(test_pwd.encode()).hexdigest()
            if user.password == md5_hash:
                password_info['formats']['md5'] = test_pwd
            
            # SHA1
            sha1_hash = hashlib.sha1(test_pwd.encode()).hexdigest()
            if user.password == sha1_hash:
                password_info['formats']['sha1'] = test_pwd
            
            # SHA256
            sha256_hash = hashlib.sha256(test_pwd.encode()).hexdigest()
            if user.password == sha256_hash:
                password_info['formats']['sha256'] = test_pwd
        
        debug_info.append(password_info)
    
    return debug_info 